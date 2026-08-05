"""Write the results Excel file with a detailed sheet and a summary sheet."""

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RESULT_COLUMNS = [
    ("url", "URL"),
    ("username", "Username"),
    ("full_name", "Full Name"),
    ("followers", "Followers"),
    ("following", "Following"),
    ("verified", "Verified"),
    ("views", "Views"),
    ("likes", "Likes"),
    ("comments", "Comments"),
    ("caption", "Caption"),
    ("post_date", "Post Date"),
    ("media_type", "Media Type"),
    ("duration_sec", "Duration (sec)"),
    ("location", "Location"),
    ("hashtags", "Hashtags"),
    ("thumbnail_url", "Thumbnail URL"),
    ("profile_url", "Profile URL"),
    ("post_link", "Post Link"),
    ("posts_count", "Posts Count"),
    ("bio", "Bio"),
    ("is_private", "Is Private"),
    ("status", "Status"),
    ("error", "Error"),
    ("last_updated", "Last Updated"),
]

COMPUTED_COLUMNS = [
    ("Engagement Rate (%)", "engagement_rate"),
    ("Like-to-Follower Ratio", "like_follower_ratio"),
    ("View-to-Follower Ratio", "view_follower_ratio"),
    ("Comment-to-Like Ratio", "comment_like_ratio"),
]

STATUS_OK = "OK"


def _build_rows(original_df: pd.DataFrame, url_rows: list, results: dict) -> pd.DataFrame:
    """Merge original data with fetched results, preserving original rows."""
    data = []
    for url_row in url_rows:
        row = original_df.loc[url_row.excel_row - 2].to_dict()
        res = results.get(url_row.url)
        if res is None:
            row["Status"] = "Pending"
            row["Error"] = ""
        else:
            for key, label in RESULT_COLUMNS:
                if key in res:
                    row[label] = res[key]
        data.append(row)

    df = pd.DataFrame(data)
    for _key, label in RESULT_COLUMNS:
        if label not in df.columns:
            df[label] = ""
    return df


def _compute_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Add computed engagement metrics to the dataframe."""

    def safe_float(val):
        try:
            if val is None or val == "" or val == "None":
                return None
            return float(val)
        except (ValueError, TypeError):
            return None

    def compute_engagement_rate(row):
        likes = safe_float(row.get("Likes"))
        comments = safe_float(row.get("Comments"))
        followers = safe_float(row.get("Followers"))
        if followers and followers > 0 and likes is not None and comments is not None:
            return round(((likes + comments) / followers) * 100, 2)
        return None

    def compute_like_follower_ratio(row):
        likes = safe_float(row.get("Likes"))
        followers = safe_float(row.get("Followers"))
        if followers and followers > 0 and likes is not None:
            return round(likes / followers, 4)
        return None

    def compute_view_follower_ratio(row):
        views = safe_float(row.get("Views"))
        followers = safe_float(row.get("Followers"))
        if followers and followers > 0 and views is not None:
            return round(views / followers, 4)
        return None

    def compute_comment_like_ratio(row):
        comments = safe_float(row.get("Comments"))
        likes = safe_float(row.get("Likes"))
        if likes and likes > 0 and comments is not None:
            return round(comments / likes, 4)
        return None

    df["Engagement Rate (%)"] = df.apply(compute_engagement_rate, axis=1)
    df["Like-to-Follower Ratio"] = df.apply(compute_like_follower_ratio, axis=1)
    df["View-to-Follower Ratio"] = df.apply(compute_view_follower_ratio, axis=1)
    df["Comment-to-Like Ratio"] = df.apply(compute_comment_like_ratio, axis=1)

    return df


def write_results(
    original_df: pd.DataFrame,
    url_rows: list,
    results: dict,
    output_path: Path,
    started_at: datetime,
    finished_at: datetime,
) -> Path:
    """Generate the results workbook and return its path."""
    df = _build_rows(original_df, url_rows, results)
    df = _compute_metrics(df)

    total = len(url_rows)
    success = sum(1 for u in url_rows if results.get(u.url, {}).get("status") == STATUS_OK)
    failed = total - success

    engagement_values = df["Engagement Rate (%)"].dropna()
    avg_engagement = round(engagement_values.mean(), 2) if len(engagement_values) > 0 else "N/A"

    summary = {
        "Metric": [
            "Total URLs",
            "Successful",
            "Failed",
            "Started At",
            "Finished At",
            "Processing Time (sec)",
            "Avg Speed (URLs/min)",
            "Avg Engagement Rate (%)",
            "Generated At",
        ],
        "Value": [
            total,
            success,
            failed,
            _fmt(started_at),
            _fmt(finished_at),
            round((finished_at - started_at).total_seconds(), 1),
            round(success / max((finished_at - started_at).total_seconds() / 60, 1e-9), 2),
            avg_engagement,
            _fmt(datetime.now(timezone.utc)),
        ],
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Results", index=False)
        pd.DataFrame(summary).to_excel(writer, sheet_name="Summary", index=False)
        _style(writer)

    return output_path


def _fmt(dt: datetime) -> str:
    return dt.astimezone().strftime("%Y-%m-%d %H:%M:%S")


def _style(writer: pd.ExcelWriter) -> None:
    wb = writer.book
    header_fill = PatternFill(start_color="171717", end_color="171717", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, col in enumerate(ws.columns, start=1):
            width = min(max(len(str(c.value or "")) for c in col[:20]) + 2, 60)
            ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)
        ws.freeze_panes = "A2"
